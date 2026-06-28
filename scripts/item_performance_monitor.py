#!/usr/bin/env python3
"""Item Performance Monitor — SKU-level sales performance classification.

Tracks sales performance for each SKU across Shopee and Lazada, classifying
items into categories:
  - Dead Stock: 0 sales in both current and previous period (28 days)
  - Falling Behind: Sales velocity dropped >30% vs previous period
  - Potential/Gaining: Sales velocity increased >30% vs previous period
  - Hero Product: Top 20% by total sales volume
  - Stable: Everything else

Usage:
    python3 scripts/item_performance_monitor.py --channel -1004450247696
    python3 scripts/item_performance_monitor.py --channel -1004450247696 --days 14 --platform both

Stdout contract:
    {"ok": true, "platform": "shopee+lazada", "channel": "...", "text": "...",
     "blocks": [...], "excel_path": "/tmp/item_performance_YYYY-MM-DD.xlsx",
     "summary": {...}}
    {"ok": false, "error": "..."}
"""

from __future__ import annotations

import argparse
import json
import math
import os
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

MALAYSIA_TZ = timezone(timedelta(hours=8))
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PYTHON_BIN = os.path.join(REPO_ROOT, ".venv", "bin", "python")
ITEM_FETCH_BATCH = 50
TEMPO_DIR = "/tmp"


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class SKURecord:
    platform: str
    item_id: str
    sku_id: str
    sku_name: str
    product_name: str
    stock: int
    current_sold: int = 0
    previous_sold: int = 0
    velocity: float = 0.0
    trend_pct: float | None = None
    stockout_days: float | None = None
    category: str = "Stable"


# ---------------------------------------------------------------------------
# Safe-run helpers
# ---------------------------------------------------------------------------

def _run_shopee(args: list[str]) -> dict:
    cmd = [PYTHON_BIN, "-m", "platform_helpers.shopee.safe_run", "--", *args]
    result = subprocess.run(cmd, capture_output=True, text=True, cwd=REPO_ROOT, timeout=300)
    if not result.stdout.strip():
        return {"ok": False, "error": result.stderr.strip() or f"empty stdout (exit={result.returncode})"}
    try:
        return json.loads(result.stdout.strip().splitlines()[-1])
    except json.JSONDecodeError as exc:
        return {"ok": False, "error": f"JSON parse error: {exc}; stdout={result.stdout[:200]!r}"}


def _run_lazada(args: list[str]) -> dict:
    cmd = [PYTHON_BIN, "-m", "platform_helpers.lazada.safe_run", "--", *args]
    result = subprocess.run(cmd, capture_output=True, text=True, cwd=REPO_ROOT, timeout=300)
    if not result.stdout.strip():
        return {"ok": False, "error": result.stderr.strip() or f"empty stdout (exit={result.returncode})"}
    try:
        return json.loads(result.stdout.strip().splitlines()[-1])
    except json.JSONDecodeError as exc:
        return {"ok": False, "error": f"JSON parse error: {exc}; stdout={result.stdout[:200]!r}"}


# ---------------------------------------------------------------------------
# Shopee data fetching
# ---------------------------------------------------------------------------

def _shopee_extract_stock(entity: dict) -> int:
    siv2 = entity.get("stock_info_v2")
    if isinstance(siv2, dict):
        summary = siv2.get("summary_info", {})
        if isinstance(summary, dict):
            return int(summary.get("total_available_stock", 0) or 0)
    return int(entity.get("stock") or 0)


def fetch_shopee_products(max_pages: int) -> dict:
    return _run_shopee(["products", "get", "--item-status", "NORMAL", "--page-size", "50", "--max-pages", str(max_pages)])


def fetch_shopee_orders(days: int, max_pages: int, time_from: int | None = None, time_to: int | None = None) -> dict:
    args = ["orders", "get", "--days", str(days), "--max-pages", str(max_pages)]
    if time_from is not None:
        args.extend(["--time-from", str(time_from)])
    if time_to is not None:
        args.extend(["--time-to", str(time_to)])
    return _run_shopee(args)


def fetch_shopee_order_items(order_sns: list[str]) -> dict:
    if not order_sns:
        return {"ok": True, "items": []}
    all_items: list[dict] = []
    for i in range(0, len(order_sns), ITEM_FETCH_BATCH):
        batch = order_sns[i : i + ITEM_FETCH_BATCH]
        result = _run_shopee(["orders", "items", "--order-sn-list", ",".join(batch)])
        if not result.get("ok"):
            return result
        all_items.extend(result.get("items", []))
    return {"ok": True, "items": all_items}


def fetch_shopee_product_detail(item_id: int) -> dict:
    return _run_shopee(["products", "get-one", "--item-id", str(item_id)])


def aggregate_shopee_sales(items_data: dict) -> dict[tuple[int, int], int]:
    sales: dict[tuple[int, int], int] = {}
    for item in items_data.get("items", []):
        item_id = item.get("item_id")
        model_id = item.get("model_id")
        if item_id is None:
            continue
        key = (int(item_id), int(model_id) if model_id is not None else 0)
        quantity = int(item.get("model_quantity_purchased") or item.get("quantity", 0))
        sales[key] = sales.get(key, 0) + quantity
    return sales


def build_shopee_records(
    products_data: dict,
    current_sales: dict[tuple[int, int], int],
    previous_sales: dict[tuple[int, int], int],
    days: int,
) -> list[SKURecord]:
    """Build SKURecords for Shopee by fetching product details for items with sales."""
    # Collect all item_ids that had sales in either period
    item_ids_with_sales: set[int] = set()
    for (item_id, _model_id) in set(current_sales.keys()) | set(previous_sales.keys()):
        item_ids_with_sales.add(item_id)

    records: list[SKURecord] = []
    for item_id in item_ids_with_sales:
        detail = fetch_shopee_product_detail(item_id)
        if not detail.get("ok"):
            continue
        has_model = detail.get("has_model", False)
        base_info = detail.get("item_base_info", {})
        product_name = base_info.get("item_name", f"Item {item_id}")

        if not has_model:
            stock = _shopee_extract_stock(base_info)
            cur = current_sales.get((item_id, 0), 0)
            prev = previous_sales.get((item_id, 0), 0)
            velocity = cur / days if days > 0 else 0
            trend = _compute_trend(cur, prev)
            stockout = stock / velocity if velocity > 0 else None
            records.append(SKURecord(
                platform="Shopee", item_id=str(item_id), sku_id="0",
                sku_name="(single)", product_name=product_name, stock=stock,
                current_sold=cur, previous_sold=prev, velocity=velocity,
                trend_pct=trend, stockout_days=stockout,
            ))
        else:
            for m in detail.get("models", []):
                model_id = m.get("model_id", 0)
                model_name = m.get("model_name", f"Model {model_id}")
                stock = _shopee_extract_stock(m)
                cur = current_sales.get((item_id, model_id), 0)
                prev = previous_sales.get((item_id, model_id), 0)
                velocity = cur / days if days > 0 else 0
                trend = _compute_trend(cur, prev)
                stockout = stock / velocity if velocity > 0 else None
                records.append(SKURecord(
                    platform="Shopee", item_id=str(item_id), sku_id=str(model_id),
                    sku_name=model_name, product_name=product_name, stock=stock,
                    current_sold=cur, previous_sold=prev, velocity=velocity,
                    trend_pct=trend, stockout_days=stockout,
                ))
    return records


# ---------------------------------------------------------------------------
# Lazada data fetching
# ---------------------------------------------------------------------------

def fetch_lazada_products(max_pages: int) -> dict:
    return _run_lazada(["products", "get", "--filter", "all", "--limit", "50", "--max-pages", str(max_pages)])


def fetch_lazada_orders(days: int, max_pages: int, created_after: str | None = None, created_before: str | None = None) -> dict:
    args = ["orders", "get", "--days", str(days), "--status", "all", "--limit", "100", "--max-pages", str(max_pages)]
    if created_after is not None:
        args.extend(["--created-after", created_after])
    if created_before is not None:
        args.extend(["--created-before", created_before])
    return _run_lazada(args)


def fetch_lazada_order_items(order_ids: list[str]) -> dict:
    if not order_ids:
        return {"ok": True, "order_items": []}
    all_items: list[dict] = []
    for i in range(0, len(order_ids), ITEM_FETCH_BATCH):
        batch = order_ids[i : i + ITEM_FETCH_BATCH]
        result = _run_lazada(["orders", "items-multiple", "--order-ids", json.dumps(batch)])
        if not result.get("ok"):
            return result
        for entry in result.get("order_items", []):
            if isinstance(entry, dict):
                items = entry.get("order_items", [])
                if isinstance(items, list):
                    all_items.extend(items)
    return {"ok": True, "order_items": all_items}


def aggregate_lazada_sales(items_data: dict) -> dict[str, int]:
    sales: dict[str, int] = {}
    for item in items_data.get("order_items", []):
        sku = item.get("sku") or item.get("seller_sku") or ""
        if not sku:
            continue
        quantity = item.get("quantity", 0) or 1
        sales[sku] = sales.get(sku, 0) + int(quantity)
    return sales


def build_lazada_records(
    products: list[dict],
    current_sales: dict[str, int],
    previous_sales: dict[str, int],
    days: int,
) -> list[SKURecord]:
    """Build SKURecords for Lazada from product listing + sales data."""
    # Build product map: item_id -> {name, skus: {seller_sku: stock}}
    product_map: dict[int, dict] = {}
    for p in products:
        item_id = p.get("item_id")
        if item_id is None:
            continue
        item_id = int(item_id)
        name = p.get("attributes", {}).get("name", f"Item {item_id}")
        skus: dict[str, int] = {}
        for sku in p.get("skus", []):
            seller_sku = sku.get("SellerSku", sku.get("seller_sku", ""))
            qty = int(sku.get("quantity", 0))
            if seller_sku:
                skus[seller_sku] = qty
        product_map[item_id] = {"name": name, "skus": skus}

    # Build reverse map: seller_sku -> item_id
    sku_to_item: dict[str, int] = {}
    for item_id, info in product_map.items():
        for seller_sku in info["skus"]:
            sku_to_item[seller_sku] = item_id

    records: list[SKURecord] = []
    all_skus = set(current_sales.keys()) | set(previous_sales.keys())
    for seller_sku in all_skus:
        item_id = sku_to_item.get(seller_sku, 0)
        pinfo = product_map.get(item_id, {"name": f"Item {item_id}", "skus": {}})
        stock = pinfo["skus"].get(seller_sku, 0)
        cur = current_sales.get(seller_sku, 0)
        prev = previous_sales.get(seller_sku, 0)
        velocity = cur / days if days > 0 else 0
        trend = _compute_trend(cur, prev)
        stockout = stock / velocity if velocity > 0 else None
        records.append(SKURecord(
            platform="Lazada", item_id=str(item_id), sku_id=seller_sku,
            sku_name=seller_sku, product_name=pinfo["name"], stock=stock,
            current_sold=cur, previous_sold=prev, velocity=velocity,
            trend_pct=trend, stockout_days=stockout,
        ))
    return records


# ---------------------------------------------------------------------------
# Trend & classification
# ---------------------------------------------------------------------------

def _compute_trend(current: int, previous: int) -> float | None:
    if previous > 0:
        return ((current - previous) / previous) * 100
    if current > 0:
        return None  # new item, no previous data — treated as "new"
    return None  # both zero


def classify_skus(records: list[SKURecord], top_pct: float = 20) -> list[SKURecord]:
    """Classify each SKU into a performance category.

    Priority: Dead Stock > Hero > Falling Behind > Potential > Stable
    """
    # Compute sales volume for hero classification (non-dead SKUs only)
    active_sold = [r.current_sold for r in records if r.current_sold > 0]
    if active_sold:
        sorted_sold = sorted(active_sold, reverse=True)
        cutoff_idx = max(1, math.ceil(len(sorted_sold) * (top_pct / 100)))
        hero_threshold = sorted_sold[cutoff_idx - 1] if cutoff_idx <= len(sorted_sold) else sorted_sold[-1]
    else:
        hero_threshold = 0

    for r in records:
        # Dead Stock: zero sales in both periods
        if r.current_sold == 0 and r.previous_sold == 0:
            r.category = "Dead Stock"
            continue

        # Hero: top N% by current period sales
        if r.current_sold >= hero_threshold and r.current_sold > 0:
            r.category = "Hero Product"
            continue

        # Trend-based classification
        if r.trend_pct is not None:
            if r.trend_pct < -30:
                r.category = "Falling Behind"
            elif r.trend_pct > 30:
                r.category = "Potential"
            else:
                r.category = "Stable"
        else:
            # New item (previous == 0, current > 0)
            if r.previous_sold == 0 and r.current_sold > 0:
                r.category = "Potential"
            else:
                r.category = "Stable"

    return records


# ---------------------------------------------------------------------------
# Output: Markdown table
# ---------------------------------------------------------------------------

CATEGORY_EMOJI = {
    "Dead Stock": "💀",
    "Falling Behind": "📉",
    "Potential": "📈",
    "Hero Product": "🦸",
    "Stable": "✅",
}

CATEGORY_ORDER = ["Hero Product", "Potential", "Falling Behind", "Dead Stock", "Stable"]


def _trend_str(trend: float | None) -> str:
    if trend is None:
        return "NEW"
    if trend >= 0:
        return f"+{trend:.0f}%"
    return f"{trend:.0f}%"


def _stockout_str(stockout: float | None) -> str:
    if stockout is None:
        return "—"
    if stockout < 1:
        return "<1d"
    return f"{stockout:.0f}d"


def build_markdown_table(records: list[SKURecord], days: int) -> str:
    now = datetime.now(MALAYSIA_TZ)
    lines: list[str] = []
    lines.append(f"📊 *Item Performance Report* — _{now.strftime('%b %-d, %Y')}_")
    lines.append(f"_Last {days} days vs previous {days} days_\n")

    grouped: dict[str, list[SKURecord]] = defaultdict(list)
    for r in records:
        grouped[r.category].append(r)

    summary_parts = []
    for cat in CATEGORY_ORDER:
        count = len(grouped.get(cat, []))
        emoji = CATEGORY_EMOJI[cat]
        summary_parts.append(f"{emoji} {cat}: {count}")
    lines.append(" | ".join(summary_parts))
    lines.append("")

    for cat in CATEGORY_ORDER:
        items = grouped.get(cat, [])
        if not items:
            continue
        emoji = CATEGORY_EMOJI[cat]
        lines.append(f"{emoji} *{cat}* ({len(items)})")

        if cat == "Dead Stock":
            # Simplified table for dead stock
            lines.append("| SKU | Product | Platform | Last Sold | Stock |")
            lines.append("|-----|---------|----------|-----------|-------|")
            for r in sorted(items, key=lambda x: x.product_name):
                last_sold = f"{days + days}d ago" if r.previous_sold == 0 else f"{days}d ago"
                lines.append(f"| {r.sku_name} | {r.product_name[:30]} | {r.platform} | {last_sold} | {r.stock} |")
        elif cat == "Stable":
            # Just show count, skip detail table
            lines.append(f"_Showing {len(items)} stable items (detail in Excel)_")
        else:
            lines.append("| SKU | Product | Platform | Sold | Trend | Stock | Stockout |")
            lines.append("|-----|---------|----------|------|-------|-------|----------|")
            for r in sorted(items, key=lambda x: x.current_sold, reverse=True):
                lines.append(
                    f"| {r.sku_name} | {r.product_name[:30]} | {r.platform} "
                    f"| {r.current_sold} | {_trend_str(r.trend_pct)} "
                    f"| {r.stock} | {_stockout_str(r.stockout_days)} |"
                )
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Output: Excel
# ---------------------------------------------------------------------------

def build_excel(records: list[SKURecord], path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # -- Summary sheet --
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Summary counts
    grouped: dict[str, list[SKURecord]] = defaultdict(list)
    for r in records:
        grouped[r.category].append(r)

    ws_summary["A1"] = "Category"
    ws_summary["B1"] = "Count"
    ws_summary["A1"].font = header_font_white
    ws_summary["B1"].font = header_font_white
    ws_summary["A1"].fill = header_fill
    ws_summary["B1"].fill = header_fill
    ws_summary["A1"].border = thin_border
    ws_summary["B1"].border = thin_border

    for i, cat in enumerate(CATEGORY_ORDER, start=2):
        ws_summary.cell(row=i, column=1, value=cat)
        ws_summary.cell(row=i, column=2, value=len(grouped.get(cat, [])))
        ws_summary.cell(row=i, column=1).border = thin_border
        ws_summary.cell(row=i, column=2).border = thin_border

    total_row = len(CATEGORY_ORDER) + 2
    ws_summary.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=2, value=len(records)).font = Font(bold=True)
    ws_summary.cell(row=total_row, column=1).border = thin_border
    ws_summary.cell(row=total_row, column=2).border = thin_border

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 12

    # -- Detail sheet --
    ws_detail = wb.create_sheet("All SKUs")
    headers = ["Category", "Platform", "Product", "SKU", "Stock", "Sold (Current)",
               "Sold (Previous)", "Velocity/Day", "Trend %", "Stockout Days"]
    for col, h in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    CATEGORY_FILLS = {
        "Dead Stock": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Falling Behind": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Potential": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Hero Product": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "Stable": None,
    }

    sorted_records = sorted(records, key=lambda r: (
        CATEGORY_ORDER.index(r.category) if r.category in CATEGORY_ORDER else 99,
        -r.current_sold,
    ))

    for row_idx, r in enumerate(sorted_records, start=2):
        values = [
            r.category, r.platform, r.product_name, r.sku_name, r.stock,
            r.current_sold, r.previous_sold,
            round(r.velocity, 2),
            _trend_str(r.trend_pct),
            round(r.stockout_days, 1) if r.stockout_days is not None else None,
        ]
        fill = CATEGORY_FILLS.get(r.category)
        for col, val in enumerate(values, 1):
            cell = ws_detail.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if fill:
                cell.fill = fill

    # Auto-fit columns
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws_detail.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(sorted_records) + 2)
        )
        ws_detail.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Freeze top row
    ws_detail.freeze_panes = "A2"

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Telegram blocks
# ---------------------------------------------------------------------------

def build_telegram_blocks(records: list[SKURecord], days: int) -> list[dict]:
    now = datetime.now(MALAYSIA_TZ)
    blocks: list[dict] = []

    header = (
        f"\U0001f4ca *Item Performance Report*\n"
        f"_{now.strftime('%a %b %-d, %Y')} \u00b7 {days}d vs {days}d_"
    )
    blocks.append({"type": "section", "text": {"type": "mrkdwn", "text": header}})

    grouped: dict[str, list[SKURecord]] = defaultdict(list)
    for r in records:
        grouped[r.category].append(r)

    # Summary line
    summary_parts = []
    for cat in CATEGORY_ORDER:
        emoji = CATEGORY_EMOJI[cat]
        count = len(grouped.get(cat, []))
        summary_parts.append(f"{emoji} {count}")
    blocks.append({"type": "section", "text": {"type": "mrkdwn", "text": " | ".join(summary_parts)}})

    # Detail sections (skip Stable — too many)
    for cat in CATEGORY_ORDER:
        if cat == "Stable":
            continue
        items = grouped.get(cat, [])
        if not items:
            continue
        emoji = CATEGORY_EMOJI[cat]
        lines = [f"{emoji} *{cat}* ({len(items)})"]

        if cat == "Dead Stock":
            for r in sorted(items, key=lambda x: x.product_name)[:8]:
                lines.append(f"  {r.sku_name} — {r.product_name[:30]} ({r.platform}) stock:{r.stock}")
            if len(items) > 8:
                lines.append(f"  _...and {len(items) - 8} more (see Excel)_")
        else:
            for r in sorted(items, key=lambda x: x.current_sold, reverse=True)[:5]:
                trend = _trend_str(r.trend_pct)
                so = _stockout_str(r.stockout_days)
                lines.append(f"  {r.sku_name} — {r.product_name[:30]} ({r.platform}) sold:{r.current_sold} {trend} stock:{r.stock} {so}")
            if len(items) > 5:
                lines.append(f"  _...and {len(items) - 5} more (see Excel)_")

        blocks.append({"type": "section", "text": {"type": "mrkdwn", "text": "\n".join(lines)}})

    # Stable count
    stable_count = len(grouped.get("Stable", []))
    if stable_count > 0:
        blocks.append({"type": "section", "text": {"type": "mrkdwn", "text": f"\u2705 *Stable:* {stable_count} items (see Excel for details)"}})

    return blocks


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description="Item Performance Monitor — SKU-level sales classification")
    parser.add_argument("--channel", required=True, help="Channel ID")
    parser.add_argument("--days", type=int, default=14, help="Lookback window per period (default: 14)")
    parser.add_argument("--platform", choices=["shopee", "lazada", "both"], default="both")
    parser.add_argument("--top-pct", type=float, default=20, help="Top N%% for hero classification")
    parser.add_argument("--max-pages", type=int, default=10)
    args = parser.parse_args()

    now = datetime.now(MALAYSIA_TZ)
    excel_path = os.path.join(TEMPO_DIR, f"item_performance_{now.strftime('%Y-%m-%d')}.xlsx")

    try:
        all_records: list[SKURecord] = []

        # --- Shopee ---
        if args.platform in ("shopee", "both"):
            print(f"[item_perf] fetching Shopee products...", file=sys.stderr)
            products_data = fetch_shopee_products(args.max_pages)
            if not products_data.get("ok"):
                print(f"[item_perf] Shopee products fetch failed: {products_data.get('error')}", file=sys.stderr)
            else:
                products = products_data.get("items", [])
                print(f"[item_perf] Shopee: {len(products)} products", file=sys.stderr)

                now_ts = int(datetime.now(timezone.utc).timestamp())
                cur_from = now_ts - args.days * 86400
                prev_from = now_ts - args.days * 2 * 86400
                prev_to = now_ts - args.days * 86400

                print(f"[item_perf] fetching Shopee orders (current: last {args.days}d)...", file=sys.stderr)
                cur_orders = fetch_shopee_orders(args.days, args.max_pages, time_from=cur_from, time_to=now_ts)
                cur_sns = [str(o.get("order_sn", "")) for o in cur_orders.get("orders", []) if o.get("order_sn")]
                cur_items = fetch_shopee_order_items(cur_sns)
                current_sales = aggregate_shopee_sales(cur_items)
                print(f"[item_perf] Shopee current: {sum(current_sales.values())} units in {len(current_sales)} SKUs", file=sys.stderr)

                print(f"[item_perf] fetching Shopee orders (previous: {args.days}d-{args.days*2}d ago)...", file=sys.stderr)
                prev_orders = fetch_shopee_orders(args.days, args.max_pages, time_from=prev_from, time_to=prev_to)
                prev_sns = [str(o.get("order_sn", "")) for o in prev_orders.get("orders", []) if o.get("order_sn")]
                prev_items = fetch_shopee_order_items(prev_sns)
                previous_sales = aggregate_shopee_sales(prev_items)
                print(f"[item_perf] Shopee previous: {sum(previous_sales.values())} units in {len(previous_sales)} SKUs", file=sys.stderr)

                records = build_shopee_records(products_data, current_sales, previous_sales, args.days)
                all_records.extend(records)
                print(f"[item_perf] Shopee: {len(records)} SKU records", file=sys.stderr)

        # --- Lazada ---
        if args.platform in ("lazada", "both"):
            print(f"[item_perf] fetching Lazada products...", file=sys.stderr)
            products_data = fetch_lazada_products(args.max_pages)
            if not products_data.get("ok"):
                print(f"[item_perf] Lazada products fetch failed: {products_data.get('error')}", file=sys.stderr)
            else:
                products = products_data.get("products", [])
                print(f"[item_perf] Lazada: {len(products)} products", file=sys.stderr)

                now_dt = datetime.now(MALAYSIA_TZ)
                cur_from_dt = now_dt - timedelta(days=args.days)
                prev_from_dt = now_dt - timedelta(days=args.days * 2)
                prev_to_dt = now_dt - timedelta(days=args.days)

                cur_from_str = cur_from_dt.strftime("%Y-%m-%d")
                cur_to_str = now_dt.strftime("%Y-%m-%d")
                prev_from_str = prev_from_dt.strftime("%Y-%m-%d")
                prev_to_str = prev_to_dt.strftime("%Y-%m-%d")

                print(f"[item_perf] fetching Lazada orders (current: {cur_from_str} to {cur_to_str})...", file=sys.stderr)
                cur_orders = fetch_lazada_orders(args.days, args.max_pages, created_after=cur_from_str, created_before=cur_to_str)
                cur_ids = [str(o.get("order_id", "")) for o in cur_orders.get("orders", []) if o.get("order_id")]
                cur_items = fetch_lazada_order_items(cur_ids)
                current_sales = aggregate_lazada_sales(cur_items)
                print(f"[item_perf] Lazada current: {sum(current_sales.values())} units in {len(current_sales)} SKUs", file=sys.stderr)

                print(f"[item_perf] fetching Lazada orders (previous: {prev_from_str} to {prev_to_str})...", file=sys.stderr)
                prev_orders = fetch_lazada_orders(args.days, args.max_pages, created_after=prev_from_str, created_before=prev_to_str)
                prev_ids = [str(o.get("order_id", "")) for o in prev_orders.get("orders", []) if o.get("order_id")]
                prev_items = fetch_lazada_order_items(prev_ids)
                previous_sales = aggregate_lazada_sales(prev_items)
                print(f"[item_perf] Lazada previous: {sum(previous_sales.values())} units in {len(previous_sales)} SKUs", file=sys.stderr)

                records = build_lazada_records(products, current_sales, previous_sales, args.days)
                all_records.extend(records)
                print(f"[item_perf] Lazada: {len(records)} SKU records", file=sys.stderr)

        if not all_records:
            payload = {
                "ok": True, "platform": args.platform, "channel": args.channel,
                "text": f"Item Performance Report ({now:%Y-%m-%d %H:%M}) — no data",
                "blocks": [{"type": "section", "text": {"type": "mrkdwn", "text": "\u2705 No SKU data found."}}],
                "summary": {"total_skus": 0},
            }
            print(json.dumps(payload, ensure_ascii=True))
            return 0

        # Classify
        classified = classify_skus(all_records, args.top_pct)

        # Build outputs
        md_table = build_markdown_table(classified, args.days)
        build_excel(classified, excel_path)
        blocks = build_telegram_blocks(classified, args.days)

        # Summary counts
        grouped: dict[str, list[SKURecord]] = defaultdict(list)
        for r in classified:
            grouped[r.category].append(r)
        summary = {cat.lower().replace(" ", "_"): len(grouped.get(cat, [])) for cat in CATEGORY_ORDER}
        summary["total_skus"] = len(classified)

        text = f"Item Performance Report ({now:%Y-%m-%d %H:%M}) — {len(classified)} SKUs"
        payload = {
            "ok": True,
            "platform": "+".join(filter(None, [
                "shopee" if args.platform in ("shopee", "both") else "",
                "lazada" if args.platform in ("lazada", "both") else "",
            ])),
            "channel": args.channel,
            "text": text,
            "blocks": blocks,
            "excel_path": excel_path,
            "summary": summary,
        }

        print(json.dumps(payload, ensure_ascii=True))
        return 0

    except Exception as exc:
        payload = {"ok": False, "error": str(exc)}
        print(json.dumps(payload, ensure_ascii=True))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
